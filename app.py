import os
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from io import BytesIO
from openpyxl.worksheet.table import Table, TableStyleInfo
import calendar
import re
import time

import pandas as pd
import streamlit as st

from dotenv import load_dotenv
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials

from youtube_api import (
    create_youtube_api,
    get_channel_info,
    get_all_channel_videos,
    split_videos_by_status,
)

from analytics import (
    create_analytics_api,
    get_period_summary,
    get_daily_channel_data,
    get_video_performance_for_day,
    get_video_analytics,
    get_previous_period,
    calculate_change,
)

from utils import (
    format_watch_time,
)

KST = ZoneInfo("Asia/Seoul")


# =========================================================
# 1. 기본 설정
# =========================================================

load_dotenv()


def get_config(name, default=None):
    value = os.getenv(name)
    if value:
        return value

    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


CLIENT_ID = get_config("GOOGLE_CLIENT_ID")
CLIENT_SECRET = get_config("GOOGLE_CLIENT_SECRET")

# 로컬에서는 localhost, 배포된 Streamlit에서는 Secrets의 주소 사용
REDIRECT_URI = get_config(
    "GOOGLE_REDIRECT_URI",
    "http://localhost:8501",
)

# HTTP 허용은 localhost 개발 때만 사용
if REDIRECT_URI.startswith("http://localhost"):
    os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
else:
    os.environ.pop("OAUTHLIB_INSECURE_TRANSPORT", None)

SCOPES = [
    "https://www.googleapis.com/auth/youtube.readonly",
    "https://www.googleapis.com/auth/yt-analytics.readonly",
]

st.set_page_config(
    page_title="Shorts Scope",
    page_icon="📊",
    layout="wide",
)

# =========================================================
# 모바일 화면 최적화
# =========================================================
st.markdown(
    """
    <style>
    @media (max-width: 768px) {
        .block-container {
            padding-top: 1.1rem !important;
            padding-left: 0.85rem !important;
            padding-right: 0.85rem !important;
            padding-bottom: 4rem !important;
        }

        h1 { font-size: 2rem !important; line-height: 1.18 !important; }
        h2 { font-size: 1.55rem !important; line-height: 1.2 !important; }
        h3 { font-size: 1.25rem !important; line-height: 1.2 !important; }

        [data-testid="stMetric"] {
            padding: 0.45rem 0.55rem !important;
            border: 1px solid rgba(128,128,128,.18);
            border-radius: 12px;
            min-height: 92px;
        }
        [data-testid="stMetricLabel"] { font-size: 0.82rem !important; }
        [data-testid="stMetricValue"] { font-size: 1.65rem !important; }
        [data-testid="stMetricDelta"] { font-size: 0.75rem !important; }

        /* 4칸 지표는 모바일에서 2 x 2 */
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(4)):not(:has(> div:nth-child(5))) {
            display: grid !important;
            grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
            gap: 0.55rem !important;
        }

        /* 5칸 현황도 너무 길어지지 않게 2열 */
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(5)):not(:has(> div:nth-child(6))) {
            display: grid !important;
            grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
            gap: 0.55rem !important;
        }

        /* 달력: 7칸을 무조건 한 줄에 유지 */
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(7)):not(:has(> div:nth-child(8))) {
            display: grid !important;
            grid-template-columns: repeat(7, minmax(0, 1fr)) !important;
            gap: 3px !important;
        }
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(7)):not(:has(> div:nth-child(8))) > div {
            min-width: 0 !important;
            width: auto !important;
        }
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(7)):not(:has(> div:nth-child(8))) button {
            min-height: 58px !important;
            padding: 3px 1px !important;
            border-radius: 7px !important;
        }
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(7)):not(:has(> div:nth-child(8))) button p {
            font-size: 0.64rem !important;
            line-height: 1.15 !important;
            white-space: pre-line !important;
        }

        /* 달력 이전/현재/다음 3칸은 가로 유지 */
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(3)):not(:has(> div:nth-child(4))) {
            flex-wrap: nowrap !important;
            gap: 0.35rem !important;
        }
        [data-testid="stHorizontalBlock"]:has(> div:nth-child(3)):not(:has(> div:nth-child(4))) > div {
            min-width: 0 !important;
        }

        .stButton > button {
            font-size: 0.82rem;
        }

        [data-testid="stVegaLiteChart"],
        [data-testid="stArrowVegaLiteChart"] {
            margin-top: -0.25rem !important;
            margin-bottom: 0.5rem !important;
        }

        hr { margin: 0.85rem 0 !important; }
        p { line-height: 1.42; }
        .stCaption, [data-testid="stCaptionContainer"] { font-size: 0.82rem !important; }

        /* 모바일에서 섹션을 조금 더 촘촘하게 */
        [data-testid="stVerticalBlock"] { gap: 0.65rem !important; }
        [data-testid="stImage"] img { border-radius: 10px !important; }
        [data-testid="stDataFrame"] { font-size: 0.78rem !important; }
        [data-baseweb="tab-list"] { gap: 0.15rem !important; }
        [data-baseweb="tab"] { padding-left: 0.45rem !important; padding-right: 0.45rem !important; }


        /* metric 제목이 ... 으로 잘리지 않도록 */
        [data-testid="stMetricLabel"] p {
            white-space: normal !important;
            overflow: visible !important;
            text-overflow: clip !important;
            line-height: 1.15 !important;
        }

        /* 긴 metric 값이 ... 으로 잘리지 않도록 */
        [data-testid="stMetricValue"] > div {
            overflow: visible !important;
            text-overflow: clip !important;
            white-space: nowrap !important;
        }
    }

    .status-two-grid {
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 0.75rem;
        margin-top: 0.6rem;
    }
    .status-card {
        border: 1px solid rgba(128,128,128,.20);
        border-radius: 12px;
        padding: 0.8rem 0.9rem;
        min-height: 88px;
    }
    .status-label {
        font-size: 0.95rem;
        margin-bottom: 0.35rem;
    }
    .status-value {
        font-size: 1.8rem;
        line-height: 1.1;
    }
    @media (max-width: 768px) {
        .status-two-grid { gap: 0.55rem; }
        .status-card {
            min-height: 92px;
            padding: 0.65rem 0.7rem;
        }
        .status-label { font-size: 0.82rem; }
        .status-value { font-size: 1.65rem; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📊 Shorts Scope")

st.write(
    "YouTube Shorts 채널의 성과를 한눈에 확인하고 분석할 수 있는 대시보드입니다."
)

st.markdown(
    """
    ### Shorts Scope에서 할 수 있는 것

    - **채널 현황 확인**: 구독자, 총 조회수, 공개 영상 수를 한눈에 확인합니다.
    - **Shorts 성과 분석**: 기간별 조회수, 시청시간, 구독자 변화와 영상별 성과를 분석합니다.
    - **성과 비교 및 진단**: 시청률, 좋아요율, 구독 전환율 등을 비교해 다음 영상 제작에 참고할 수 있습니다.

    Shorts Scope는 사용자가 직접 Google 계정을 연결한 경우에만
    YouTube의 **읽기 전용 데이터**를 불러옵니다.
    Shorts Scope는 사용자의 YouTube 콘텐츠를 생성, 수정 또는 삭제하지 않습니다.
    """
)

st.link_button(
    "🔒 개인정보처리방침",
    "https://goldmoon123.github.io/alssulcut-dashboard/privacy.html",
)

st.divider()


# =========================================================
# 2. OAuth 설정
# =========================================================

if not CLIENT_ID or not CLIENT_SECRET:

    st.error(
        "Google OAuth 설정을 찾을 수 없습니다."
    )

    st.stop()


client_config = {
    "web": {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "auth_uri":
            "https://accounts.google.com/o/oauth2/auth",
        "token_uri":
            "https://oauth2.googleapis.com/token",
        "redirect_uris": [
            REDIRECT_URI
        ],
    }
}


def create_flow():

    return Flow.from_client_config(
        client_config,
        scopes=SCOPES,
        redirect_uri=REDIRECT_URI,
        autogenerate_code_verifier=False,
    )


# =========================================================
# 3. OAuth Callback
# =========================================================

code = st.query_params.get("code")
oauth_error = st.query_params.get("error")


if oauth_error:

    st.error(
        f"Google 로그인이 실패했습니다: "
        f"{oauth_error}"
    )

    if st.button("🔄 다시 로그인"):

        st.query_params.clear()

        st.rerun()

    st.stop()


if (
    code
    and "credentials"
    not in st.session_state
):

    try:

        flow = create_flow()

        flow.fetch_token(
            code=code
        )

        credentials = flow.credentials

        st.session_state[
            "credentials"
        ] = {

            "token":
                credentials.token,

            "refresh_token":
                credentials.refresh_token,

            "token_uri":
                credentials.token_uri,

            "client_id":
                credentials.client_id,

            "client_secret":
                credentials.client_secret,

            "scopes":
                list(
                    credentials.scopes
                    or SCOPES
                ),
        }

        st.query_params.clear()

        st.rerun()

    except Exception as e:

        st.error(
            "Google 로그인 처리 중 "
            "오류가 발생했습니다."
        )

        st.code(
            str(e)
        )

        st.stop()


# =========================================================
# 4. 로그인 전
# =========================================================

if (
    "credentials"
    not in st.session_state
):

    flow = create_flow()

    authorization_url, _ = (
        flow.authorization_url(
            access_type="offline",
            prompt="consent",
            include_granted_scopes="true",
        )
    )

    st.subheader(
        "🔗 YouTube 채널 연결"
    )

    st.write(
        "Google 계정을 연결하면 본인 YouTube 채널의 읽기 전용 데이터를 불러와 "
        "채널 및 Shorts 성과 분석을 시작합니다."
    )

    st.caption(
        "연결 전에도 위에서 Shorts Scope의 목적과 개인정보처리방침을 확인할 수 있습니다."
    )

    st.link_button(
        "🔴 Google로 YouTube 연결",
        authorization_url,
        type="primary",
    )

    st.stop()


# =========================================================
# 5. Credentials 복원
# =========================================================

saved = st.session_state[
    "credentials"
]

credentials = Credentials(

    token=
        saved["token"],

    refresh_token=
        saved.get(
            "refresh_token"
        ),

    token_uri=
        saved["token_uri"],

    client_id=
        saved["client_id"],

    client_secret=
        saved["client_secret"],

    scopes=
        saved.get(
            "scopes",
            SCOPES
        ),
)


# =========================================================
# 6. API 연결
# =========================================================

try:

    youtube = create_youtube_api(
        credentials
    )

    yt_analytics = (
        create_analytics_api(
            credentials
        )
    )

except Exception as e:

    st.error(
        "YouTube API 연결 실패"
    )

    st.code(
        str(e)
    )

    st.stop()


# =========================================================
# 7. 채널 + 영상 정보
# =========================================================

try:

    channel_info = (
        get_channel_info(
            youtube
        )
    )

    if not channel_info:

        st.error(
            "YouTube 채널을 찾지 못했습니다."
        )

        st.stop()

    videos = (
        get_all_channel_videos(
            youtube,
            channel_info
        )
    )

    split = (
        split_videos_by_status(
            videos
        )
    )

    public_videos = (
        split["public"]
    )

    scheduled_videos = (
        split["scheduled"]
    )

    private_videos = (
        split["private"]
    )

    unlisted_videos = (
        split["unlisted"]
    )

except Exception as e:

    st.error(
        "YouTube 데이터를 가져오지 못했습니다."
    )

    st.code(
        str(e)
    )

    st.stop()


# =========================================================
# 8. 영상별 Analytics
# =========================================================

public_video_ids = [

    video["video_id"]

    for video in public_videos
]


try:

    video_analytics = (
        get_video_analytics(
            yt_analytics,
            public_video_ids
        )
    )

except Exception as e:

    video_analytics = {}

    st.warning(
        "영상별 Analytics 일부를 "
        "불러오지 못했습니다."
    )

    with st.expander(
        "오류 내용"
    ):

        st.code(
            str(e)
        )


# =========================================================
# 9. 영상 데이터 + Analytics 합치기
# =========================================================

for video in videos:

    data = video_analytics.get(
        video["video_id"],
        {}
    )

    video["watch_minutes"] = float(
        data.get(
            "watch_minutes",
            0
        )
    )

    video["avg_duration"] = float(
        data.get(
            "average_view_duration",
            0
        )
    )

    video["avg_percentage"] = float(
        data.get(
            "average_view_percentage",
            0
        )
    )

    video["shares"] = int(
        data.get(
            "shares",
            0
        )
    )

    video["subs_gained"] = int(
        data.get(
            "subscribers_gained",
            0
        )
    )

    video["subs_lost"] = int(
        data.get(
            "subscribers_lost",
            0
        )
    )

    video["net_subs"] = int(
        data.get(
            "net_subscribers",
            0
        )
    )

    # 비율 지표
    views_for_rate = max(video["views"], 1)
    video["like_rate"] = (video["likes"] / views_for_rate) * 100
    video["comment_rate"] = (video["comments"] / views_for_rate) * 100
    video["sub_conversion_rate"] = (video["net_subs"] / views_for_rate) * 100


# =========================================================
# 9-1. 자동 성과 점수 V1
# =========================================================

# 조회수 100회 이상 + Analytics가 실제로 잡힌 공개 영상만 평가
eligible_videos = [
    video
    for video in public_videos
    if video.get("views", 0) >= 100
    and video.get("video_id") in video_analytics
]


def percentile_scores(items, key):
    if not items:
        return {}

    values = sorted(
        (item.get(key, 0), item["video_id"])
        for item in items
    )
    total = len(values)

    if total == 1:
        return {values[0][1]: 100.0}

    result = {}
    for index, (_, video_id) in enumerate(values):
        result[video_id] = (index / (total - 1)) * 100

    return result


views_pct = percentile_scores(eligible_videos, "views")
retention_pct = percentile_scores(eligible_videos, "avg_percentage")
likes_pct = percentile_scores(eligible_videos, "like_rate")
subs_pct = percentile_scores(eligible_videos, "sub_conversion_rate")


for video in public_videos:

    if video not in eligible_videos:
        video["performance_score"] = None
        video["performance_grade"] = "⏳ 데이터 부족"
        video["diagnosis_strengths"] = []
        video["diagnosis_weaknesses"] = []
        continue

    score = (
        views_pct.get(video["video_id"], 0) * 0.30
        + retention_pct.get(video["video_id"], 0) * 0.35
        + likes_pct.get(video["video_id"], 0) * 0.15
        + subs_pct.get(video["video_id"], 0) * 0.20
    )

    video["performance_score"] = round(score)

    if score >= 80:
        video["performance_grade"] = "🔥 매우 좋음"
    elif score >= 65:
        video["performance_grade"] = "🟢 좋음"
    elif score >= 45:
        video["performance_grade"] = "🟡 보통"
    else:
        video["performance_grade"] = "🔴 개선 필요"

    strengths = []
    weaknesses = []

    if retention_pct.get(video["video_id"], 0) >= 70:
        strengths.append("시청 유지")
    elif retention_pct.get(video["video_id"], 0) <= 30:
        weaknesses.append("시청 유지")

    if subs_pct.get(video["video_id"], 0) >= 70:
        strengths.append("구독 전환")
    elif subs_pct.get(video["video_id"], 0) <= 30:
        weaknesses.append("구독 전환")

    if likes_pct.get(video["video_id"], 0) >= 70:
        strengths.append("좋아요 반응")
    elif likes_pct.get(video["video_id"], 0) <= 30:
        weaknesses.append("좋아요 반응")

    if views_pct.get(video["video_id"], 0) >= 70:
        strengths.append("조회수")
    elif views_pct.get(video["video_id"], 0) <= 30:
        weaknesses.append("조회수")

    video["diagnosis_strengths"] = strengths
    video["diagnosis_weaknesses"] = weaknesses


# =========================================================
# 10. 채널 상단
# =========================================================

st.success(
    "✅ YouTube 채널 연결 완료!"
)

st.subheader(
    f"📺 {channel_info['channel_name']}"
)


c1, c2, c3 = st.columns(3)

c1.metric(
    "구독자",
    f"{channel_info['subscribers']:,}명"
)

c2.metric(
    "총 조회수",
    f"{channel_info['total_views']:,}회"
)

c3.metric(
    "공개 영상",
    f"{channel_info['public_video_count']:,}개"
)

st.divider()


# =========================================================
# 11. 영상 현황
# =========================================================

st.subheader(
    "🎬 영상 현황"
)

c1, c2, c3 = st.columns(3)

c1.metric(
    "전체",
    f"{len(videos)}개"
)

c2.metric(
    "🟢 공개",
    f"{len(public_videos)}개"
)

c3.metric(
    "🟡 예약",
    f"{len(scheduled_videos)}개"
)

st.markdown(
    f"""
    <div class="status-two-grid">
        <div class="status-card">
            <div class="status-label">🔒 비공개</div>
            <div class="status-value">{len(private_videos)}개</div>
        </div>
        <div class="status-card">
            <div class="status-label">🔵 일부공개</div>
            <div class="status-value">{len(unlisted_videos)}개</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.caption(
    "※ 예약·비공개·일부공개 영상은 "
    "성과 평균과 순위에서 제외됩니다."
)

st.divider()


# =========================================================
# 12. 날짜 / 기간 분석
# =========================================================

st.header(
    "📅 날짜 / 기간 분석"
)

st.write(
    "기간을 선택하면 해당 기간 동안 "
    "채널이 얼마나 성장했는지 확인할 수 있습니다."
)


# ---------------------------------------------------------
# 기간 선택 + 분석하기 버튼
# ---------------------------------------------------------

today = datetime.now(KST).date()

# 처음 들어왔을 때는 최근 7일을 기본 적용
if "applied_period_option" not in st.session_state:
    st.session_state.applied_period_option = "최근 7일"
    st.session_state.applied_start_date = today - timedelta(days=7)
    st.session_state.applied_end_date = today - timedelta(days=1)

period_option = st.radio(
    "분석 기간",
    ["오늘", "최근 7일", "최근 28일", "직접 선택"],
    horizontal=True,
    index=1,
    key="period_option_input",
)

selected_range = None
if period_option == "직접 선택":
    selected_range = st.date_input(
        "날짜 범위 선택",
        value=(today - timedelta(days=6), today),
        key="period_custom_range",
    )

if period_option == "오늘":
    candidate_start = today
    candidate_end = today
elif period_option == "최근 7일":
    candidate_end = today - timedelta(days=1)
    candidate_start = candidate_end - timedelta(days=6)
elif period_option == "최근 28일":
    candidate_end = today - timedelta(days=1)
    candidate_start = candidate_end - timedelta(days=27)
else:
    if isinstance(selected_range, tuple) and len(selected_range) == 2:
        candidate_start, candidate_end = selected_range
    elif selected_range:
        candidate_start = selected_range
        candidate_end = selected_range
    else:
        candidate_start = today
        candidate_end = today

if st.button(
    "🔍 분석하기",
    type="primary",
    use_container_width=True,
    key="apply_period_button",
):
    st.session_state.applied_period_option = period_option
    st.session_state.applied_start_date = candidate_start
    st.session_state.applied_end_date = candidate_end

period_option = st.session_state.applied_period_option
start_date = st.session_state.applied_start_date
end_date = st.session_state.applied_end_date

st.caption(
    f"적용된 분석 기간: {start_date} ~ {end_date}"
)

if period_option == "오늘":
    st.info(
        "⏳ 오늘 데이터는 YouTube Analytics에서 아직 집계 중일 수 있습니다. "
        "오늘 수치가 0으로 보여도 실제 조회가 없는 뜻은 아닐 수 있으며, "
        "확정 데이터는 시간이 지나면서 반영됩니다."
    )


# =========================================================
# 13. 현재 기간 / 이전 기간
# =========================================================

try:

    current_summary = (
        get_period_summary(
            yt_analytics,
            start_date,
            end_date,
        )
    )

    (
        previous_start,
        previous_end,
    ) = get_previous_period(
        start_date,
        end_date,
    )

    previous_summary = (
        get_period_summary(
            yt_analytics,
            previous_start,
            previous_end,
        )
    )

except Exception as e:

    st.error(
        "기간별 Analytics 데이터를 "
        "가져오지 못했습니다."
    )

    st.code(
        str(e)
    )

    st.stop()


# =========================================================
# 14. 변화율 함수
# =========================================================

def change_text(
    current,
    previous,
):

    change = calculate_change(
        current,
        previous
    )

    if change is None:
        return f"이전 기간 {previous:,.0f}"

    return f"{change:+.1f}%"


# =========================================================
# 15. 기간 성과 카드
# =========================================================

st.subheader(
    "📊 선택 기간 성과"
)

c1, c2, c3, c4 = (
    st.columns(4)
)


c1.metric(
    "조회수",
    f"{current_summary['views']:,}회",
    change_text(
        current_summary["views"],
        previous_summary["views"],
    ),
)


c2.metric(
    "순구독자",
    (
        f"{current_summary['net_subscribers']:+,}명"
    ),
    change_text(
        current_summary[
            "net_subscribers"
        ],
        previous_summary[
            "net_subscribers"
        ],
    ),
)


c3.metric(
    "시청시간",
    format_watch_time(
        current_summary[
            "watch_minutes"
        ]
    ),
    change_text(
        current_summary[
            "watch_minutes"
        ],
        previous_summary[
            "watch_minutes"
        ],
    ),
)


c4.metric(
    "좋아요",
    f"{current_summary['likes']:,}개",
    change_text(
        current_summary["likes"],
        previous_summary["likes"],
    ),
)


c1, c2, c3, c4 = (
    st.columns(4)
)


c1.metric(
    "구독자 획득",
    (
        f"+{current_summary['subscribers_gained']:,}명"
    )
)


c2.metric(
    "구독자 이탈",
    (
        f"-{current_summary['subscribers_lost']:,}명"
    )
)


c3.metric(
    "댓글",
    f"{current_summary['comments']:,}개"
)


c4.metric(
    "공유",
    f"{current_summary['shares']:,}회"
)


st.caption(
    f"비교 기간: "
    f"{previous_start} ~ {previous_end}"
)

st.divider()


# =========================================================
# 16. 일별 데이터
# =========================================================

try:

    daily_data = (
        get_daily_channel_data(
            yt_analytics,
            start_date,
            end_date,
        )
    )

except Exception as e:

    daily_data = []

    st.warning(
        "일별 데이터를 불러오지 못했습니다."
    )

    st.code(
        str(e)
    )


if daily_data:

    daily_df = pd.DataFrame(daily_data)
    daily_df["date"] = pd.to_datetime(daily_df["date"])

    # 데이터가 없는 날짜도 0으로 채움
    full_dates = pd.date_range(start=start_date, end=end_date, freq="D")
    daily_df = daily_df.set_index("date").reindex(full_dates).fillna(0)
    daily_df.index.name = "날짜"
    daily_df.index = [dt.strftime("%m/%d") for dt in daily_df.index]

    st.subheader("📈 일별 조회수")
    st.line_chart(daily_df[["views"]], use_container_width=True, height=260)

    st.subheader("👤 일별 순구독자")
    st.bar_chart(daily_df[["net_subscribers"]], use_container_width=True, height=260)

else:
    st.info("선택한 기간에 일별 Analytics 데이터가 없습니다.")

st.divider()


def add_excel_table(ws, table_name):
    """Excel 실제 표(Table) + 필터/정렬."""
    if ws.max_row < 2 or ws.max_column < 1:
        return
    safe_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)
    if not safe_name or safe_name[0].isdigit():
        safe_name = "T_" + safe_name
    table = Table(displayName=safe_name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def call_with_retry(func, *args, attempts=2, delay=0.35, **kwargs):
    """YouTube API의 일시적 5xx 오류를 짧게 재시도합니다."""
    last_error = None
    for attempt in range(attempts):
        try:
            return func(*args, **kwargs)
        except Exception as exc:
            last_error = exc
            status = getattr(getattr(exc, "resp", None), "status", None)
            if status is not None and int(status) < 500:
                raise
            if attempt < attempts - 1:
                time.sleep(delay)
    if last_error:
        raise last_error


# =========================================================
# 17. 월간 성과 달력
# =========================================================

st.header("🗓️ 월간 성과 달력")

# 달력은 항상 현재 날짜가 속한 달부터 시작
if "calendar_month" not in st.session_state:
    st.session_state.calendar_month = today.replace(day=1)

if "calendar_selected_day" not in st.session_state:
    st.session_state.calendar_selected_day = today

calendar_month = st.session_state.calendar_month
left, center, right = st.columns([1, 3, 1])

with left:
    if st.button("◀ 이전 달", use_container_width=True):
        st.session_state.calendar_month = (calendar_month - timedelta(days=1)).replace(day=1)
        st.rerun()

with center:
    st.markdown(
        f"<h3 style='text-align:center;'>{calendar_month.year}년 {calendar_month.month}월</h3>",
        unsafe_allow_html=True,
    )

with right:
    next_month = (calendar_month.replace(day=28) + timedelta(days=4)).replace(day=1)
    if st.button(
        "다음 달 ▶",
        use_container_width=True,
        disabled=next_month > today.replace(day=1),
    ):
        st.session_state.calendar_month = next_month
        st.rerun()

month_start = calendar_month
last_day = calendar.monthrange(calendar_month.year, calendar_month.month)[1]
month_end = min(date(calendar_month.year, calendar_month.month, last_day), today)

try:
    month_daily_data = get_daily_channel_data(yt_analytics, month_start, month_end)
except Exception:
    month_daily_data = []

month_lookup = {item["date"]: item for item in month_daily_data}


def make_monthly_excel(month_daily_rows, month_value, all_public_videos):
    """
    달력에서 선택한 달을 실제 분석용 엑셀로 내보냅니다.

    1시트: 월간 일별 성과
    2시트: 그달 업로드 영상
    3시트: 월간 요약
    """
    output = BytesIO()

    # -----------------------------
    # 그달 업로드 영상 정리
    # -----------------------------
    uploaded_rows = []
    upload_by_date = {}

    for video in all_public_videos:
        published_raw = video.get("published_raw")
        if not published_raw:
            continue

        try:
            published_dt = datetime.fromisoformat(
                published_raw.replace("Z", "+00:00")
            ).astimezone(KST)
        except Exception:
            continue

        if (
            published_dt.year == month_value.year
            and published_dt.month == month_value.month
        ):
            date_key = published_dt.strftime("%Y-%m-%d")
            upload_by_date.setdefault(date_key, []).append(video.get("title", ""))

            uploaded_rows.append({
                "업로드일(KST)": published_dt.strftime("%Y-%m-%d %H:%M"),
                "제목": video.get("title", ""),
                "조회수": int(video.get("views", 0)),
                "좋아요": int(video.get("likes", 0)),
                "댓글": int(video.get("comments", 0)),
                "공유": int(video.get("shares", 0)),
                "평균 시청시간(초)": round(float(video.get("avg_duration", 0)), 1),
                "평균 시청률(%)": round(float(video.get("avg_percentage", 0)), 1),
                "순구독자": int(video.get("net_subs", 0)),
                "좋아요율(%)": round(float(video.get("like_rate", 0)), 2),
                "구독전환율(%)": round(float(video.get("sub_conversion_rate", 0)), 3),
                "성과점수": (
                    video.get("performance_score")
                    if video.get("performance_score") is not None
                    else "평가 보류"
                ),
                "성과등급": video.get("performance_grade", "-"),
            })

    month_upload_df = pd.DataFrame(uploaded_rows)

    # -----------------------------
    # 날짜별 Analytics 정리
    # 데이터가 없는 날짜도 0으로 포함
    # -----------------------------
    daily_lookup = {
        str(item.get("date", "")): item
        for item in month_daily_rows
    }

    last_day = calendar.monthrange(month_value.year, month_value.month)[1]
    calendar_last_date = date(month_value.year, month_value.month, last_day)

    if month_value.year == today.year and month_value.month == today.month:
        export_last_date = today
    else:
        export_last_date = calendar_last_date

    weekday_ko = ["월", "화", "수", "목", "금", "토", "일"]
    daily_export = []

    current_date = date(month_value.year, month_value.month, 1)

    while current_date <= export_last_date:
        date_key = current_date.strftime("%Y-%m-%d")
        item = daily_lookup.get(date_key, {})

        watch_minutes = float(item.get("watch_minutes", 0) or 0)
        uploaded_titles = upload_by_date.get(date_key, [])

        daily_export.append({
            "날짜": date_key,
            "요일": weekday_ko[current_date.weekday()],
            "조회수": int(item.get("views", 0) or 0),
            "시청시간(시간)": round(watch_minutes / 60, 2),
            "좋아요": int(item.get("likes", 0) or 0),
            "댓글": int(item.get("comments", 0) or 0),
            "공유": int(item.get("shares", 0) or 0),
            "구독자 획득": int(item.get("subscribers_gained", 0) or 0),
            "구독자 이탈": int(item.get("subscribers_lost", 0) or 0),
            "순구독자": int(item.get("net_subscribers", 0) or 0),
            "업로드 영상 수": len(uploaded_titles),
            "업로드 영상": " / ".join(uploaded_titles),
        })

        current_date += timedelta(days=1)

    month_daily_df = pd.DataFrame(daily_export)

    # 맨 아래 합계 행
    if not month_daily_df.empty:
        total_row = {
            "날짜": "합계",
            "요일": "",
            "조회수": int(month_daily_df["조회수"].sum()),
            "시청시간(시간)": round(float(month_daily_df["시청시간(시간)"].sum()), 2),
            "좋아요": int(month_daily_df["좋아요"].sum()),
            "댓글": int(month_daily_df["댓글"].sum()),
            "공유": int(month_daily_df["공유"].sum()),
            "구독자 획득": int(month_daily_df["구독자 획득"].sum()),
            "구독자 이탈": int(month_daily_df["구독자 이탈"].sum()),
            "순구독자": int(month_daily_df["순구독자"].sum()),
            "업로드 영상 수": int(month_daily_df["업로드 영상 수"].sum()),
            "업로드 영상": "",
        }
        month_daily_df = pd.concat(
            [month_daily_df, pd.DataFrame([total_row])],
            ignore_index=True,
        )

    # -----------------------------
    # 월간 요약
    # -----------------------------
    data_only_df = month_daily_df[month_daily_df["날짜"] != "합계"].copy()

    summary_df = pd.DataFrame([
        ["채널명", channel_info.get("channel_name", channel_info.get("title", ""))],
        ["대상 월", f"{month_value.year}-{month_value.month:02d}"],
        ["월 조회수", int(data_only_df["조회수"].sum()) if not data_only_df.empty else 0],
        ["월 순구독자", int(data_only_df["순구독자"].sum()) if not data_only_df.empty else 0],
        ["월 시청시간(시간)", round(float(data_only_df["시청시간(시간)"].sum()), 2) if not data_only_df.empty else 0],
        ["월 좋아요", int(data_only_df["좋아요"].sum()) if not data_only_df.empty else 0],
        ["월 댓글", int(data_only_df["댓글"].sum()) if not data_only_df.empty else 0],
        ["월 공유", int(data_only_df["공유"].sum()) if not data_only_df.empty else 0],
        ["그달 업로드 영상", len(month_upload_df)],
        ["생성 시각", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S KST")],
    ], columns=["항목", "값"])

    # -----------------------------
    # Excel 출력
    # -----------------------------
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 열자마자 상세 데이터가 먼저 보이게 함
        month_daily_df.to_excel(writer, sheet_name="월간 일별 성과", index=False)
        month_upload_df.to_excel(writer, sheet_name="그달 업로드 영상", index=False)
        summary_df.to_excel(writer, sheet_name="월간 요약", index=False)

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header_map = {
                cell.value: cell.column_letter
                for cell in ws[1]
                if cell.value is not None
            }

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        max_length = max(max_length, cell_length)
                    except Exception:
                        pass

                ws.column_dimensions[column_letter].width = min(
                    max(max_length + 3, 11),
                    48,
                )

            # 날짜는 무조건 문자열로
            for header in ["날짜", "업로드일(KST)"]:
                if header in header_map:
                    col = header_map[header]
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col}{row}"].number_format = "@"

            if "날짜" in header_map:
                ws.column_dimensions[header_map["날짜"]].width = 14

            if "업로드일(KST)" in header_map:
                ws.column_dimensions[header_map["업로드일(KST)"]].width = 21

            if "제목" in header_map:
                ws.column_dimensions[header_map["제목"]].width = 42

            if "업로드 영상" in header_map:
                ws.column_dimensions[header_map["업로드 영상"]].width = 48

        for idx, sheet_name in enumerate(writer.book.sheetnames, start=1):
            add_excel_table(writer.book[sheet_name], f"MonthlyTable_{idx}")

    output.seek(0)
    return output.getvalue()


# 공개 영상을 올린 날짜(KST)
upload_dates = set()
for video in public_videos:
    published_raw = video.get("published_raw")
    if not published_raw:
        continue
    try:
        published_dt = datetime.fromisoformat(
            published_raw.replace("Z", "+00:00")
        ).astimezone(KST)
        upload_dates.add(published_dt.date())
    except Exception:
        pass

weekday_names = ["월", "화", "수", "목", "금", "토", "일"]
for column, name in zip(st.columns(7), weekday_names):
    column.markdown(f"<div style='text-align:center;'><b>{name}</b></div>", unsafe_allow_html=True)

for week in calendar.monthcalendar(calendar_month.year, calendar_month.month):
    columns = st.columns(7)
    for day_index, day_number in enumerate(week):
        if day_number == 0:
            columns[day_index].write("")
            continue

        current_day = date(calendar_month.year, calendar_month.month, day_number)
        day_key = current_day.strftime("%Y-%m-%d")
        has_analytics_row = day_key in month_lookup
        day_data = month_lookup.get(day_key, {})
        views = int(day_data.get("views", 0))
        net_subscribers = int(day_data.get("net_subscribers", 0))

        if current_day > today:
            columns[day_index].button(
                f"{day_number}일\n\n-",
                key=f"future_{current_day}",
                disabled=True,
                use_container_width=True,
            )
            continue

        if views >= 10000:
            view_text = f"{views / 10000:.1f}만"
        elif views >= 1000:
            view_text = f"{views / 1000:.1f}천"
        else:
            view_text = f"{views:,}"

        subscriber_text = f"\n👤 {net_subscribers:+d}" if net_subscribers != 0 else ""
        upload_mark = " 🎬" if current_day in upload_dates else ""

        # 최근 2일은 Analytics 행 자체가 없으면 0으로 단정하지 않음
        is_recent_calendar_day = 0 <= (today - current_day).days <= 2
        if is_recent_calendar_day and not has_analytics_row:
            label = f"{day_number}일{upload_mark}\n\n⏳ 집계 중"
        elif views == 0 and net_subscribers == 0:
            label = f"{day_number}일{upload_mark}"
        else:
            label = f"{day_number}일{upload_mark}\n\n👁 {view_text}{subscriber_text}"

        if columns[day_index].button(
            label,
            key=f"calendar_{current_day}",
            use_container_width=True,
        ):
            st.session_state.calendar_selected_day = current_day
            st.rerun()

st.caption("날짜를 누르면 아래에서 그날의 상세 성과를 확인할 수 있습니다.")

st.download_button(
    f"📥 {calendar_month.year}년 {calendar_month.month}월 성과 엑셀",
    data=make_monthly_excel(month_daily_data, calendar_month, public_videos),
    file_name=f"shorts_monthly_{calendar_month.year}_{calendar_month.month:02d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    key=f"monthly_excel_{calendar_month.year}_{calendar_month.month}",
)

st.divider()

# =========================================================
# 18. 특정 날짜 선택
# =========================================================

st.header(
    "🔎 하루 자세히 보기"
)

if "applied_detail_day" not in st.session_state:
    st.session_state.applied_detail_day = st.session_state.get(
        "calendar_selected_day",
        end_date,
    )

detail_day_input = st.date_input(
    "확인할 날짜",
    value=st.session_state.get("calendar_selected_day", end_date),
    min_value=date(2005, 1, 1),
    max_value=today,
    key="detail_day_input",
)
st.session_state.calendar_selected_day = detail_day_input

if st.button(
    "🔍 조회하기",
    type="primary",
    use_container_width=True,
    key="apply_detail_day_button",
):
    st.session_state.applied_detail_day = detail_day_input

selected_day = st.session_state.applied_detail_day
st.caption(f"현재 조회 중인 날짜: {selected_day}")

# 최근 날짜는 YouTube Analytics의 일별 집계가 아직 완료되지 않았을 수 있음
detail_age_days = (today - selected_day).days
is_recent_detail = 0 <= detail_age_days <= 2
if is_recent_detail:
    st.warning(
        "⏳ 최근 날짜의 YouTube Analytics 데이터는 아직 집계 중일 수 있습니다. "
        "아래 일별 수치가 0으로 보여도 실제 0이라고 단정할 수 없습니다. "
        "영상 아래의 '현재 조회수'는 누적값이고, 여기의 '그날 조회수'는 날짜별 Analytics 값입니다."
    )


try:

    day_summary = (
        get_period_summary(
            yt_analytics,
            selected_day,
            selected_day,
        )
    )

except Exception as e:

    st.error(
        "선택 날짜 데이터를 "
        "가져오지 못했습니다."
    )

    st.code(
        str(e)
    )

    day_summary = None


if day_summary:

    st.subheader(
        f"📅 {selected_day}"
    )

    c1, c2, c3, c4 = (
        st.columns(4)
    )

    c1.metric(
        "그날 조회수",
        f"+{day_summary['views']:,}회"
    )

    c2.metric(
        "그날 순구독자",
        (
            f"{day_summary['net_subscribers']:+,}명"
        )
    )

    c3.metric(
        "그날 시청시간",
        format_watch_time(
            day_summary[
                "watch_minutes"
            ]
        )
    )

    c4.metric(
        "그날 좋아요",
        f"+{day_summary['likes']:,}개"
    )

    if is_recent_detail and (
        day_summary.get("views", 0) == 0
        and day_summary.get("watch_minutes", 0) == 0
        and day_summary.get("likes", 0) == 0
        and day_summary.get("net_subscribers", 0) == 0
    ):
        st.info(
            "📌 이 날짜의 Analytics 값은 아직 미집계일 가능성이 있습니다. "
            "0을 확정 성과로 해석하지 마세요."
        )


    # =====================================================
    # 그날 영상별 성과
    # =====================================================

    try:

        day_video_data = (
            get_video_performance_for_day(
                yt_analytics,
                selected_day,
            )
        )

    except Exception as e:

        day_video_data = []

        st.warning(
            "그날의 영상별 데이터를 "
            "불러오지 못했습니다."
        )

        st.code(
            str(e)
        )


    video_lookup = {

        video["video_id"]:
            video

        for video in videos
    }

    # 당일 업로드 영상 vs 기존 영상 조회수 기여도
    uploaded_ids_for_day = set()
    for video in public_videos:
        published_raw = video.get("published_raw")
        if not published_raw:
            continue
        try:
            published_dt = datetime.fromisoformat(
                published_raw.replace("Z", "+00:00")
            ).astimezone(KST)
            if published_dt.date() == selected_day:
                uploaded_ids_for_day.add(video["video_id"])
        except Exception:
            pass

    new_video_views = sum(
        item["views"] for item in day_video_data
        if item["video_id"] in uploaded_ids_for_day
    )
    old_video_views = max(day_summary["views"] - new_video_views, 0)

    st.subheader("🧩 그날 조회수 구성")

    day_total_views = max(day_summary["views"], 0)

    if day_total_views > 0:
        new_video_share = (new_video_views / day_total_views) * 100
        old_video_share = (old_video_views / day_total_views) * 100
    else:
        new_video_share = 0.0
        old_video_share = 0.0

    cc1, cc2 = st.columns(2)

    cc1.metric(
        "당일 업로드 영상",
        f"{new_video_views:,}회",
        f"{new_video_share:.1f}% 기여",
    )

    cc2.metric(
        "기존 영상",
        f"{old_video_views:,}회",
        f"{old_video_share:.1f}% 기여",
    )

    if is_recent_detail and day_total_views == 0:
        st.info(
            "📌 현재 0회는 '확정 0회'가 아니라 Analytics 일별 집계 대기일 가능성이 있습니다. "
            "오늘 업로드 영상의 실제 누적 조회수는 아래 '그날 업로드한 영상'의 현재 조회수를 확인하세요."
        )


    if day_video_data:

        st.subheader(
            "🔥 그날 조회수를 만든 영상"
        )

        for rank, performance in enumerate(
            day_video_data[:5],
            start=1,
        ):

            video = video_lookup.get(
                performance["video_id"]
            )

            if not video:
                continue

            col_img, col_info = (
                st.columns(
                    [1, 5]
                )
            )

            with col_img:

                if video[
                    "thumbnail"
                ]:

                    st.image(
                        video[
                            "thumbnail"
                        ],
                        width=140
                    )


            with col_info:

                st.markdown(
                    f"### {rank}위 · "
                    f"{video['title']}"
                )

                st.write(
                    f"👁️ 그날 "
                    f"+{performance['views']:,}회"
                    f"  |  "
                    f"👍 "
                    f"+{performance['likes']:,}"
                    f"  |  "
                    f"👤 "
                    f"{performance['net_subscribers']:+d}"
                )

        st.divider()


# =========================================================
# 18. 그날 업로드한 영상
# =========================================================

uploaded_that_day = []


for video in public_videos:

    published_raw = (
        video.get(
            "published_raw"
        )
    )

    if not published_raw:
        continue

    try:

        published_dt = (
            datetime.fromisoformat(
                published_raw.replace(
                    "Z",
                    "+00:00"
                )
            )
        )

        published_dt = published_dt.astimezone(KST)

        if (
            published_dt.date()
            == selected_day
        ):

            uploaded_that_day.append(
                video
            )

    except Exception:
        pass


st.subheader(
    "🎬 그날 업로드한 영상"
)


if uploaded_that_day:

    for video in uploaded_that_day:

        col_img, col_info = (
            st.columns(
                [1, 5]
            )
        )

        with col_img:

            if video[
                "thumbnail"
            ]:

                st.image(
                    video[
                        "thumbnail"
                    ],
                    width=140
                )


        with col_info:

            st.markdown(
                f"**{video['title']}**"
            )

            st.write(
                f"현재 조회수 "
                f"{video['views']:,}회"
            )

            st.write(
                f"영상 길이 "
                f"{video['duration']}"
            )

else:

    st.write(
        "이날 업로드한 영상이 없습니다."
    )


st.divider()


# =========================================================
# 19. 공개 영상 전체 성과
# =========================================================

st.header(
    "🎯 전체 공개 영상 분석"
)


if public_videos:

    total_views = sum(
        video["views"]
        for video in public_videos
    )

    average_views = (
        total_views
        / len(public_videos)
    )

    best_video = max(
        public_videos,
        key=lambda x: x["views"]
    )

    total_watch_minutes = sum(
        video["watch_minutes"]
        for video in public_videos
    )

    total_net_subscribers = sum(
        video["net_subs"]
        for video in public_videos
    )


    c1, c2, c3, c4 = (
        st.columns(4)
    )

    c1.metric(
        "평균 조회수",
        f"{average_views:,.0f}회"
    )

    c2.metric(
        "최고 조회수",
        f"{best_video['views']:,}회"
    )

    c3.metric(
        "총 시청시간",
        format_watch_time(
            total_watch_minutes
        )
    )

    c4.metric(
        "공개 영상 순구독자 합계",
        f"{total_net_subscribers:+,}명"
    )


st.divider()


# =========================================================
# 20. 자동 성과 리포트 V6
# =========================================================

st.header("🧠 자동 성과 리포트 V6.2")
st.caption(f"🕒 데이터 조회 시각: {datetime.now(KST).strftime('%Y-%m-%d %H:%M KST')} · 최근 날짜의 Analytics는 지연될 수 있습니다.")
st.caption(
    "원인을 추측하지 않고 실제 데이터와 내 채널 기준선만 비교합니다. "
    "조회수 100회 이상 + 영상별 Analytics가 집계된 공개 영상만 분석합니다."
)

report_videos = [
    v for v in public_videos
    if v.get("views", 0) >= 100 and v.get("video_id") in video_analytics
]
st.caption(
    f"분석 대상: 공개 영상 {len(public_videos)}개 중 {len(report_videos)}개 · "
    f"마지막 조회: {datetime.now(KST).strftime('%Y-%m-%d %H:%M KST')}"
)

# -----------------------------
# 채널 추세
# -----------------------------
st.markdown("### 📈 채널 추세")

trend_option = st.selectbox(
    "추세 분석 기간",
    ["최근 7일", "최근 14일", "최근 28일", "직접 선택"],
    key="trend_period_option_v61",
)

trend_last_day = today - timedelta(days=1)

if trend_option == "최근 7일":
    trend_end = trend_last_day
    trend_start = trend_end - timedelta(days=6)
elif trend_option == "최근 14일":
    trend_end = trend_last_day
    trend_start = trend_end - timedelta(days=13)
elif trend_option == "최근 28일":
    trend_end = trend_last_day
    trend_start = trend_end - timedelta(days=27)
else:
    tc1, tc2 = st.columns(2)
    with tc1:
        trend_start = st.date_input(
            "추세 시작일",
            value=trend_last_day - timedelta(days=6),
            max_value=trend_last_day,
            key="trend_start_v61",
        )
    with tc2:
        trend_end = st.date_input(
            "추세 종료일",
            value=trend_last_day,
            max_value=trend_last_day,
            key="trend_end_v61",
        )

if trend_start > trend_end:
    st.warning("시작일이 종료일보다 늦어 종료일 기준으로 맞췄습니다.")
    trend_start = trend_end

trend_days = (trend_end - trend_start).days + 1
trend_prev_end = trend_start - timedelta(days=1)
trend_prev_start = trend_prev_end - timedelta(days=trend_days - 1)

def _count_uploads(start_d, end_d):
    count = 0
    for _v in public_videos:
        _raw = _v.get("published_raw")
        if not _raw:
            continue
        try:
            _dt = datetime.fromisoformat(_raw.replace("Z", "+00:00")).astimezone(KST)
            if start_d <= _dt.date() <= end_d:
                count += 1
        except Exception:
            pass
    return count

def _trend_change(cur, prev):
    if prev == 0:
        return "계산 불가" if cur != 0 else "0%"
    return f"{((cur-prev)/abs(prev))*100:+.1f}%"

try:
    trend_now = get_period_summary(yt_analytics, trend_start, trend_end)
    trend_prev = get_period_summary(yt_analytics, trend_prev_start, trend_prev_end)
    trend_daily_rows = get_daily_channel_data(yt_analytics, trend_start, trend_end)
except Exception as exc:
    trend_now = None
    trend_prev = None
    trend_daily_rows = []
    st.warning("채널 추세 데이터를 일부 불러오지 못했습니다.")
    with st.expander("오류 내용"):
        st.code(str(exc))

if trend_now and trend_prev:
    now_uploads = _count_uploads(trend_start, trend_end)
    prev_uploads = _count_uploads(trend_prev_start, trend_prev_end)

    st.caption(
        f"현재 {trend_start} ~ {trend_end} ↔ 이전 {trend_prev_start} ~ {trend_prev_end} · 오늘 제외"
    )

    trend_table = pd.DataFrame([
        ["조회수", f"{trend_now['views']:,}회", f"{trend_prev['views']:,}회",
         _trend_change(trend_now["views"], trend_prev["views"])],
        ["시청시간", f"{trend_now['watch_minutes']/60:,.1f}시간",
         f"{trend_prev['watch_minutes']/60:,.1f}시간",
         _trend_change(trend_now["watch_minutes"], trend_prev["watch_minutes"])],
        ["순구독자", f"{trend_now['net_subscribers']:+,}명",
         f"{trend_prev['net_subscribers']:+,}명",
         f"{trend_now['net_subscribers']-trend_prev['net_subscribers']:+,}명"],
        ["업로드", f"{now_uploads:,}개", f"{prev_uploads:,}개",
         f"{now_uploads-prev_uploads:+,}개"],
    ], columns=["지표", "현재 기간", "이전 기간", "변화"])

    st.dataframe(trend_table, hide_index=True, use_container_width=True)

    if prev_uploads == 0:
        st.warning(
            "⚠️ 이전 기간 업로드가 0개입니다. 조회수 증가율이 커 보여도 "
            "콘텐츠 자체의 성과가 같은 비율로 개선됐다고 볼 수는 없습니다."
        )
    elif trend_prev["views"] == 0:
        st.warning(
            "⚠️ 이전 기간 조회수가 0회라 변화율 비교가 의미 없습니다."
        )

    if now_uploads > 0 and prev_uploads > 0:
        now_per_upload = trend_now["views"] / now_uploads
        prev_per_upload = trend_prev["views"] / prev_uploads
        tc1, tc2 = st.columns(2)
        tc1.metric("조회수 ÷ 업로드 수 (참고)", f"{now_per_upload:,.0f}회")
        tc2.metric("이전 기간", f"{prev_per_upload:,.0f}회")
        st.caption(
            "※ 이 값에는 기존 영상 조회수도 포함됩니다. 신규 영상 1편의 실제 평균 조회수는 아닙니다."
        )
    else:
        st.caption("※ 두 기간 모두 업로드가 있을 때만 '조회수 ÷ 업로드 수'를 표시합니다.")

    if trend_daily_rows:
        trend_df = pd.DataFrame(trend_daily_rows)
        if not trend_df.empty and "date" in trend_df.columns and "views" in trend_df.columns:
            trend_df["date"] = pd.to_datetime(trend_df["date"])
            trend_df = trend_df.set_index("date")
            st.markdown("#### 일별 조회수 흐름")
            st.line_chart(trend_df[["views"]], use_container_width=True, height=240)

st.divider()

if report_videos:
    base_views = sum(v["views"] for v in report_videos) / len(report_videos)
    median_views = float(pd.Series([v["views"] for v in report_videos]).median())
    base_ret = sum(v["avg_percentage"] for v in report_videos) / len(report_videos)
    base_like = sum(v["like_rate"] for v in report_videos) / len(report_videos)
    base_sub = sum(v["sub_conversion_rate"] for v in report_videos) / len(report_videos)

    st.markdown("### 📊 내 채널 기준선")
    b1,b2,b3,b4,b5=st.columns(5)
    b1.metric("평균 조회수", f"{base_views:,.0f}회")
    b2.metric("중앙 조회수", f"{median_views:,.0f}회")
    b3.metric("평균 시청률", f"{base_ret:.1f}%")
    b4.metric("평균 좋아요율", f"{base_like:.2f}%")
    b5.metric("평균 구독전환율", f"{base_sub:.3f}%")
    st.caption("※ 위 기준선은 현재 분석 가능한 내 영상들의 비교값이며 YouTube 공식 기준이 아닙니다.")

    # -----------------------------
    # 영상 나이 / 성장 참고
    # -----------------------------
    st.markdown("### ⏱️ 영상 나이 · 성장 참고")
    st.caption(
        "현재 누적 조회수를 업로드 후 경과일로 나눈 참고값입니다. "
        "과거 같은 시점의 실제 조회수를 복원한 값은 아니므로 '동일 시점 성장곡선'과는 다릅니다."
    )

    age_rows = []
    for _v in report_videos:
        _raw = _v.get("published_raw")
        if not _raw:
            continue
        try:
            _dt = datetime.fromisoformat(_raw.replace("Z", "+00:00")).astimezone(KST)
            _published_date = _dt.date()
            _age_days = max((today - _published_date).days, 0)
        except Exception:
            continue

        # 업로드 당일은 1일로 나눠 0 division 방지.
        _days_for_rate = max(_age_days, 1)
        _views_per_day = _v.get("views", 0) / _days_for_rate

        age_rows.append({
            "video_id": _v.get("video_id"),
            "제목": _v.get("title", ""),
            "업로드일": _published_date.strftime("%Y.%m.%d"),
            "경과일": _age_days,
            "현재 조회수": int(_v.get("views", 0)),
            "일평균 조회수(참고)": _views_per_day,
        })

    if age_rows:
        _speed_values = sorted(row["일평균 조회수(참고)"] for row in age_rows)

        def _speed_percentile(value):
            if len(_speed_values) == 1:
                return 100.0
            below_or_equal = sum(1 for x in _speed_values if x <= value)
            return (below_or_equal - 1) / (len(_speed_values) - 1) * 100

        for row in age_rows:
            row["성장속도 백분위(참고)"] = _speed_percentile(row["일평균 조회수(참고)"])

        age_rows = sorted(
            age_rows,
            key=lambda r: r["성장속도 백분위(참고)"],
            reverse=True,
        )

        age_display = pd.DataFrame([
            {
                "영상": row["제목"],
                "업로드": row["업로드일"],
                "경과": f"{row['경과일']}일",
                "현재 조회수": f"{row['현재 조회수']:,}회",
                "조회수 ÷ 경과일": f"{row['일평균 조회수(참고)']:,.0f}회/일",
                "채널 내 위치": f"상위 {max(1, round(100 - row['성장속도 백분위(참고)']))}%",
            }
            for row in age_rows
        ])

        st.dataframe(
            age_display,
            hide_index=True,
            use_container_width=True,
        )
        st.info(
            "📌 이 표의 '상위 %'는 현재 누적 조회수를 경과일로 나눈 값끼리 비교한 참고치입니다. "
            "업로드 후 24시간·3일·7일의 실제 성과를 비교하려면 영상별 일별 기록을 따로 저장하거나 "
            "YouTube Analytics의 영상별 일자 데이터를 추가로 수집해야 합니다."
        )
    else:
        st.info("업로드 날짜를 확인할 수 있는 분석 대상 영상이 없습니다.")

    st.markdown("### 🔬 영상별 데이터 비교")
    st.caption("기본 화면은 압축되어 있습니다. 영상을 누르면 채널 기준선과 실제 차이를 확인할 수 있습니다.")

    ordered=sorted(report_videos,key=lambda v:v.get("views",0),reverse=True)

    def pct_diff(value, base):
        return "비교 불가" if base == 0 else f"{((value-base)/abs(base))*100:+.1f}%"

    def pp_diff(value, base, digits):
        diff = value - base
        threshold = {
            1: 1.0,      # 시청률: ±1.0%p 이내는 비슷
            2: 0.05,     # 좋아요율: ±0.05%p 이내는 비슷
            3: 0.005,    # 구독전환율: ±0.005%p 이내는 비슷
        }.get(digits, 0)
        if abs(diff) <= threshold:
            return "≈ 비슷"
        return f"{diff:+.{digits}f}%p"

    def compare_level(value, base, metric_name):
        if metric_name == "조회수":
            if base == 0:
                return "similar"
            diff_ratio = abs(value - base) / abs(base)
            if diff_ratio <= 0.05:
                return "similar"
        elif metric_name == "시청률":
            if abs(value - base) <= 1.0:
                return "similar"
        elif metric_name == "좋아요율":
            if abs(value - base) <= 0.05:
                return "similar"
        elif metric_name == "구독전환율":
            if abs(value - base) <= 0.005:
                return "similar"

        return "high" if value > base else "low"

    for rank,v in enumerate(ordered,start=1):
        published_text = "업로드일 확인 불가"
        age_text = ""
        raw = v.get("published_raw")
        if raw:
            try:
                published_dt = datetime.fromisoformat(raw.replace("Z", "+00:00")).astimezone(KST)
                published_date = published_dt.date()
                age_days = max((today - published_date).days, 0)
                published_text = published_date.strftime("%Y.%m.%d")
                age_text = f" · 업로드 후 {age_days}일"
            except Exception:
                pass

        summary=(f"📅 {published_text}{age_text} | "
                 f"조회수 {v['views']:,} · 시청률 {v['avg_percentage']:.1f}% · "
                 f"좋아요 {v['like_rate']:.2f}% · 구독 {v['sub_conversion_rate']:.3f}%")
        with st.expander(f"{rank}. {v['title']}  |  {summary}"):
            rows=[
                ["조회수",f"{v['views']:,}회",f"{base_views:,.0f}회",pct_diff(v["views"],base_views)],
                ["평균 시청률",f"{v['avg_percentage']:.1f}%",f"{base_ret:.1f}%",pp_diff(v["avg_percentage"],base_ret,1)],
                ["좋아요율",f"{v['like_rate']:.2f}%",f"{base_like:.2f}%",pp_diff(v["like_rate"],base_like,2)],
                ["구독전환율",f"{v['sub_conversion_rate']:.3f}%",f"{base_sub:.3f}%",pp_diff(v["sub_conversion_rate"],base_sub,3)],
            ]
            st.dataframe(pd.DataFrame(rows,columns=["지표","이 영상","채널 기준선","차이"]),
                         hide_index=True,use_container_width=True)

            high=[]; low=[]; similar=[]
            for name,val,base in [
                ("조회수",v["views"],base_views),("시청률",v["avg_percentage"],base_ret),
                ("좋아요율",v["like_rate"],base_like),("구독전환율",v["sub_conversion_rate"],base_sub)]:
                level = compare_level(val, base, name)
                if level == "high":
                    high.append(name)
                elif level == "low":
                    low.append(name)
                else:
                    similar.append(name)

            st.markdown("**데이터에서 확인되는 점**")
            parts=[]
            if high: parts.append("기준선보다 높음: "+", ".join(high))
            if similar: parts.append("비슷한 수준: "+", ".join(similar))
            if low: parts.append("기준선보다 낮음: "+", ".join(low))
            st.write(" · ".join(parts) if parts else "채널 기준선과 비슷한 수준입니다.")

            st.markdown("**다음 테스트**")
            st.write(
                "이 영상과 비슷한 소재·길이·전개 중 한 요소를 유지한 영상을 추가로 테스트해 "
                "같은 성과가 반복되는지 확인해보세요. 현재 수치만으로 원인을 단정하지 않습니다."
            )
else:
    st.info("아직 비교 가능한 영상이 없습니다.")


# =========================================================
# 21. 전체 영상 분석
# =========================================================

st.subheader(
    "🔎 전체 영상 분석"
)


def format_video_upload_kst(video):
    """
    Excel/표에서 날짜가 #### 또는 이상한 숫자로 보이지 않도록
    YouTube 원본 업로드 시각을 한국시간 문자열로 고정합니다.
    """
    published_raw = video.get("published_raw")

    if published_raw:
        try:
            published_dt = datetime.fromisoformat(
                published_raw.replace("Z", "+00:00")
            ).astimezone(KST)
            return published_dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass

    published_value = video.get("published", "")
    if published_value is None:
        return ""

    return str(published_value)


table_data = []


for video in videos:

    has_analytics = (
        video["video_id"]
        in video_analytics
    )

    table_data.append({

        "상태":
            video["status"],

        "제목":
            video["title"],

        "조회수":
            video["views"],

        "좋아요":
            video["likes"],

        "댓글":
            video["comments"],

        "공유":
            video["shares"],

        "좋아요율":
            f"{video['like_rate']:.2f}%",

        "댓글률":
            f"{video['comment_rate']:.2f}%",

        "구독전환율":
            f"{video['sub_conversion_rate']:.3f}%",

        "성과점수":
            (
                (
                    video.get("performance_score")
                    if video.get("performance_score") is not None
                    else "평가 보류"
                )
                if video["status"] == "🟢 공개"
                else "-"
            ),

        "성과등급":
            (
                video.get("performance_grade", "-")
                if video["status"] == "🟢 공개"
                else "-"
            ),

        "평균 시청":
            (
                f"{video['avg_duration']:.1f}초"
                if has_analytics
                else "-"
            ),

        "평균 시청률":
            (
                f"{video['avg_percentage']:.1f}%"
                if has_analytics
                else "-"
            ),

        "구독자":
            (
                video["net_subs"]
                if has_analytics
                else 0
            ),

        "길이":
            video["duration"],

        "업로드":
            format_video_upload_kst(video),

        "예약 공개":
            video["scheduled"],
    })


df = pd.DataFrame(
    table_data
)


# =========================================================
# 22-1. 영상 검색 / 조회수 필터
# =========================================================

st.caption(
    "조건을 정한 뒤 검색 버튼을 눌러 결과를 확인하세요. "
    "검색 전에는 기존 결과가 그대로 유지됩니다."
)

# 조회수 조건은 바꾸면 입력칸만 바뀌고, 실제 결과는 검색 버튼을 눌러야 적용됩니다.
filter_mode_input = st.radio(
    "조회수 조건",
    ["전체", "이상", "이하", "범위"],
    horizontal=True,
    key="views_filter_mode_input",
)

filter_min_input = 0
filter_max_input = 0
filter_views_input = 5000

if filter_mode_input in ["이상", "이하"]:
    filter_views_input = st.number_input(
        "기준 조회수",
        min_value=0,
        value=5000,
        step=500,
        key="views_filter_value_input",
    )
elif filter_mode_input == "범위":
    range_c1, range_c2 = st.columns(2)
    with range_c1:
        filter_min_input = st.number_input(
            "최소 조회수",
            min_value=0,
            value=1000,
            step=500,
            key="views_filter_min_input",
        )
    with range_c2:
        filter_max_input = st.number_input(
            "최대 조회수",
            min_value=0,
            value=10000,
            step=500,
            key="views_filter_max_input",
        )

filter_c1, filter_c2 = st.columns(2)
with filter_c1:
    status_options = ["전체"] + sorted(df["상태"].dropna().astype(str).unique().tolist())
    filter_status_input = st.selectbox(
        "공개 상태",
        status_options,
        key="video_status_filter_input",
    )
with filter_c2:
    title_query_input = st.text_input(
        "제목 검색",
        placeholder="예: 비버, 화산, 교통사고",
        key="video_title_filter_input",
    )

sort_option_input = st.selectbox(
    "정렬",
    [
        "조회수 높은 순",
        "조회수 낮은 순",
        "최신 업로드 순",
        "오래된 업로드 순",
        "평균 시청률 높은 순",
        "구독 증가 높은 순",
    ],
    key="video_sort_input",
)

if "applied_video_filter" not in st.session_state:
    st.session_state.applied_video_filter = {
        "mode": "전체",
        "views": 5000,
        "min_views": 1000,
        "max_views": 10000,
        "status": "전체",
        "title": "",
        "sort": "조회수 높은 순",
    }

if st.button(
    "🔍 검색",
    type="primary",
    use_container_width=True,
    key="apply_video_filter_button",
):
    st.session_state.applied_video_filter = {
        "mode": filter_mode_input,
        "views": int(filter_views_input),
        "min_views": int(filter_min_input),
        "max_views": int(filter_max_input),
        "status": filter_status_input,
        "title": title_query_input.strip(),
        "sort": sort_option_input,
    }

applied_filter = st.session_state.applied_video_filter
filtered_df = df.copy()

if applied_filter["mode"] == "이상":
    filtered_df = filtered_df[filtered_df["조회수"] >= applied_filter["views"]]
elif applied_filter["mode"] == "이하":
    filtered_df = filtered_df[filtered_df["조회수"] <= applied_filter["views"]]
elif applied_filter["mode"] == "범위":
    low = min(applied_filter["min_views"], applied_filter["max_views"])
    high = max(applied_filter["min_views"], applied_filter["max_views"])
    filtered_df = filtered_df[
        (filtered_df["조회수"] >= low) & (filtered_df["조회수"] <= high)
    ]

if applied_filter["status"] != "전체":
    filtered_df = filtered_df[filtered_df["상태"] == applied_filter["status"]]

if applied_filter["title"]:
    filtered_df = filtered_df[
        filtered_df["제목"].astype(str).str.contains(
            applied_filter["title"], case=False, na=False
        )
    ]

# 검색 결과 정렬
sort_name = applied_filter["sort"]
if sort_name == "조회수 높은 순":
    filtered_df = filtered_df.sort_values("조회수", ascending=False)
elif sort_name == "조회수 낮은 순":
    filtered_df = filtered_df.sort_values("조회수", ascending=True)
elif sort_name == "평균 시청률 높은 순":
    filtered_df = filtered_df.assign(
        _sort_pct=pd.to_numeric(
            filtered_df["평균 시청률"].astype(str).str.replace("%", "", regex=False),
            errors="coerce",
        )
    ).sort_values("_sort_pct", ascending=False).drop(columns=["_sort_pct"])
elif sort_name == "구독 증가 높은 순":
    filtered_df = filtered_df.sort_values("구독자", ascending=False)
elif sort_name in ["최신 업로드 순", "오래된 업로드 순"]:
    filtered_df = filtered_df.assign(
        _sort_date=pd.to_datetime(filtered_df["업로드"], errors="coerce")
    ).sort_values(
        "_sort_date",
        ascending=(sort_name == "오래된 업로드 순"),
    ).drop(columns=["_sort_date"])

result_count = len(filtered_df)
result_avg_views = int(filtered_df["조회수"].mean()) if result_count else 0
result_max_views = int(filtered_df["조회수"].max()) if result_count else 0

fc1, fc2, fc3 = st.columns(3)
fc1.metric("검색 영상", f"{result_count:,}개")
fc2.metric("평균 조회수", f"{result_avg_views:,}회")
fc3.metric("최고 조회수", f"{result_max_views:,}회")

st.caption(
    f"현재 적용: 조회수 {applied_filter['mode']} · "
    f"상태 {applied_filter['status']} · 정렬 {applied_filter['sort']}"
)

with st.expander(
    f"📋 검색 결과 보기 ({result_count:,}개)",
    expanded=(0 < result_count <= 10),
):
    if filtered_df.empty:
        st.info("조건에 맞는 영상이 없습니다.")
    else:
        st.dataframe(
            filtered_df,
            use_container_width=True,
            hide_index=True,
            height=min(420, 90 + (len(filtered_df) * 34)),
        )


# =========================================================
# 22-2. 엑셀 다운로드
# =========================================================

def make_excel_file(primary_df, primary_sheet_name):
    """
    전체 영상/검색 결과 엑셀.
    날짜·길이를 엑셀이 임의의 날짜/시간 형식으로 바꾸지 않도록 문자열로 내보냅니다.
    """
    output = BytesIO()

    export_df = primary_df.copy()

    # Excel이 업로드 날짜/영상 길이를 자동 날짜·시간으로 오인하지 않게 문자열 고정
    for column_name in ["업로드", "예약 공개", "길이"]:
        if column_name in export_df.columns:
            export_df[column_name] = export_df[column_name].fillna("").astype(str)

    summary_rows = [
        ["채널명", channel_info.get("channel_name", channel_info.get("title", ""))],
        ["구독자", channel_info.get("subscribers", 0)],
        ["채널 총 조회수", channel_info.get("total_views", channel_info.get("views", 0))],
        ["전체 감지 영상", len(videos)],
        ["공개 영상", len(public_videos)],
        ["예약 영상", len(scheduled_videos)],
        ["내보낸 영상", len(export_df)],
        ["생성 시각", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S KST")],
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["항목", "값"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=primary_sheet_name, index=False)
        summary_df.to_excel(writer, sheet_name="채널 요약", index=False)

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        max_length = max(max_length, cell_length)
                    except Exception:
                        pass

                ws.column_dimensions[column_letter].width = min(
                    max(max_length + 3, 11),
                    48,
                )

            # 영상 데이터 시트에서 자주 잘리던 열은 최소 폭 보장
            if sheet_name == primary_sheet_name:
                header_map = {
                    cell.value: cell.column_letter
                    for cell in ws[1]
                    if cell.value is not None
                }

                if "업로드" in header_map:
                    ws.column_dimensions[header_map["업로드"]].width = 21

                if "예약 공개" in header_map:
                    ws.column_dimensions[header_map["예약 공개"]].width = 21

                if "길이" in header_map:
                    ws.column_dimensions[header_map["길이"]].width = 12

                if "제목" in header_map:
                    ws.column_dimensions[header_map["제목"]].width = 42

                # 날짜/시간 열은 '텍스트'로 고정
                for header in ["업로드", "예약 공개", "길이"]:
                    if header in header_map:
                        col = header_map[header]
                        for row in range(2, ws.max_row + 1):
                            ws[f"{col}{row}"].number_format = "@"

        for idx, sheet_name in enumerate(writer.book.sheetnames, start=1):
            add_excel_table(writer.book[sheet_name], f"VideoTable_{idx}")

    output.seek(0)
    return output.getvalue()

st.markdown("### 📥 엑셀 다운로드")
download_c1, download_c2 = st.columns(2)

with download_c1:
    st.download_button(
        "📥 전체 영상 엑셀",
        data=make_excel_file(df, "전체 영상"),
        file_name=f"shorts_all_videos_{today.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with download_c2:
    st.download_button(
        f"📥 검색 결과 엑셀 ({result_count}개)",
        data=make_excel_file(filtered_df, "검색 결과") if result_count else b"",
        file_name=f"shorts_filtered_videos_{today.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        disabled=result_count == 0,
    )

with st.expander("📋 전체 영상 데이터 보기", expanded=False):
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        height=340,
    )




# =========================================================
# 22. 영상 TOP 랭킹
# =========================================================

with st.expander("🏆 공개 영상 TOP 랭킹 보기", expanded=False):
    rank_limit = st.radio(
        "표시 개수",
        [5, 10, 20],
        horizontal=True,
        format_func=lambda n: f"TOP {n}",
        key="ranking_display_count",
    )

    rank_tab1, rank_tab2, rank_tab3, rank_tab4 = st.tabs(
        ["👁️ 조회수", "📊 시청률", "👤 구독전환", "👍 좋아요율"]
    )

    def render_ranked_videos(ranked, metric_name, metric_formatter):
        if not ranked:
            st.info("표시할 공개 영상이 없습니다.")
            return

        for rank, video in enumerate(ranked[:rank_limit], start=1):
            col_img, col_info = st.columns([1, 5])
            with col_img:
                if video["thumbnail"]:
                    st.image(video["thumbnail"], width=125)
            with col_info:
                st.markdown(f"**{rank}위 · {video['title']}**")
                score_value = video.get("performance_score")

                if score_value is None:
                    score_text = "⏳ 데이터 부족"
                else:
                    score_text = (
                        f"🚦 {score_value}점 · "
                        f"{video.get('performance_grade', '-')}"
                    )

                st.write(
                    f"**{metric_name}: {metric_formatter(video)}**  |  "
                    f"{score_text}"
                )
                st.write(
                    f"👁️ {video['views']:,}회  |  "
                    f"👍 {video['likes']:,} ({video['like_rate']:.2f}%)  |  "
                    f"💬 {video['comments']:,} ({video['comment_rate']:.2f}%)  |  "
                    f"👤 {video['net_subs']:+d} ({video['sub_conversion_rate']:.3f}%)"
                )
                if video["video_id"] in video_analytics:
                    st.write(
                        f"⏱️ 평균 시청 {video['avg_duration']:.1f}초  |  "
                        f"📊 평균 시청률 {video['avg_percentage']:.1f}%  |  "
                        f"🎞️ {video['duration']}"
                    )
                st.link_button(
                    "▶️ YouTube에서 보기",
                    "https://www.youtube.com/watch?v=" + video["video_id"],
                    key=f"rank_{metric_name}_{video['video_id']}"
                )
            st.divider()

    with rank_tab1:
        render_ranked_videos(
            sorted(public_videos, key=lambda x: x["views"], reverse=True),
            "조회수", lambda v: f"{v['views']:,}회"
        )

    with rank_tab2:
        render_ranked_videos(
            sorted(public_videos, key=lambda x: x["avg_percentage"], reverse=True),
            "평균 시청률", lambda v: f"{v['avg_percentage']:.1f}%"
        )

    with rank_tab3:
        render_ranked_videos(
            sorted(public_videos, key=lambda x: x["sub_conversion_rate"], reverse=True),
            "구독전환율", lambda v: f"{v['sub_conversion_rate']:.3f}%"
        )

    with rank_tab4:
        render_ranked_videos(
            sorted(public_videos, key=lambda x: x["like_rate"], reverse=True),
            "좋아요율", lambda v: f"{v['like_rate']:.2f}%"
        )



# =========================================================
# 23. 예약 영상
# =========================================================

if scheduled_videos:

    st.divider()

    st.subheader(
        "🟡 예약 영상"
    )


    for video in scheduled_videos:

        col_img, col_info = (
            st.columns(
                [1, 5]
            )
        )


        with col_img:

            if video[
                "thumbnail"
            ]:

                st.image(
                    video[
                        "thumbnail"
                    ],
                    width=140
                )


        with col_info:

            st.markdown(
                f"**{video['title']}**"
            )

            st.write(
                f"공개 예정: "
                f"{video['scheduled']}"
            )

            st.write(
                f"영상 길이: "
                f"{video['duration']}"
            )


# =========================================================
# 24. 연결 해제
# =========================================================

st.divider()


if st.button(
    "🔓 YouTube 연결 해제"
):

    st.session_state.clear()

    st.query_params.clear()

    st.rerun()
